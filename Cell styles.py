Python 3.10.4 (tags/v3.10.4:9d38120, Mar 23 2022, 23:13:41) [MSC v.1929 64 bit (AMD64)] on win32
Type "help", "copyright", "credits" or "license()" for more information.
import openpyxl

die(openpyxl.styles)
Traceback (most recent call last):
  File "<pyshell#2>", line 1, in <module>
    die(openpyxl.styles)
NameError: name 'die' is not defined. Did you mean: 'dir'?
dir(openpyxl.styles)
['Alignment', 'Border', 'Color', 'DEFAULT_FONT', 'Fill', 'Font', 'GradientFill', 'NamedStyle', 'NumberFormatDescriptor', 'PatternFill', 'Protection', 'Side', '__builtins__', '__cached__', '__doc__', '__file__', '__loader__', '__name__', '__package__', '__path__', '__spec__', 'alignment', 'borders', 'builtins', 'cell_style', 'colors', 'differential', 'fills', 'fonts', 'is_builtin', 'is_date_format', 'named_styles', 'numbers', 'protection', 'proxy', 'styleable', 'stylesheet', 'table']
from openpyxl.styles import *

workbook = openpyxl.load_workbook("E:\Employees.xlsx")
sheet = workbook['EmployeeData']
cell = sheet['B8']


font = Font(color = colors.BLUE, bold = True,italic = True)

cell.font = font

fill = PatternFill(fill_type = 'solid',bgColor = 'F7FE2E')


cell.fill = fill


border = Border (left =Side(border_style = 'double',color = '322FEC'),right = Side(border_style = 'double',color = '322FEC'), top = Side(border_style = 'double',color = '322FEC')bottom = Side(border_style = 'double'color = '322FEC')))
SyntaxError: unmatched ')'
border = Border (left =Side(border_style = 'double',color = '322FEC'),right = Side(border_style = 'double',color = '322FEC'), top = Side(border_style = 'double',color = '322FEC')bottom = Side(border_style = 'double'color = '322FEC')))border = Border (left =Side(border_style = 'double',color = '322FEC'),right = Side(border_style = 'double',color = '322FEC'), top = Side(border_style = 'double',color = '322FEC')bottom = Side(border_style = 'double'color = '322FEC')
                                                                                                                                                                                                                                                           
SyntaxError: unmatched ')'
border = Border (left =Side(border_style = 'double',color = '322FEC'),right = Side(border_style = 'double',color = '322FEC'), top = Side(border_style = 'double',color = '322FEC')bottom = Side(border_style = 'double'color = '322FEC')
                 
SyntaxError: invalid syntax. Perhaps you forgot a comma?
border = Border (left =Side(border_style = 'double',color = '322FEC'),right = Side(border_style = 'double',color = '322FEC'), top = Side(border_style = 'double',color = '322FEC')bottom = Side(border_style = 'double',color = '322FEC')
                 
SyntaxError: invalid syntax. Perhaps you forgot a comma?
SyntaxError: invalid syntax. Perhaps you forgot a comma?
                 
SyntaxError: invalid syntax


border = Border(left = Side(border_style = 'double', color = '322FEC'), right = Side(border_style = 'double', color = '322FEC'), top = Side(border_style = 'double', color = '322FEC'), bottom = Side(border_style = 'double', color = '322FEC'))
                 
cell.border = border
                 
align = Alignment(horizontal = 'left')
                 
cell.alignment = align
                 
workbook.save("E:\Employees.xlsx")
                 
