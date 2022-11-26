Python 3.10.4 (tags/v3.10.4:9d38120, Mar 23 2022, 23:13:41) [MSC v.1929 64 bit (AMD64)] on win32
Type "help", "copyright", "credits" or "license()" for more information.
import openpyxl

workbook = openpyxl.load_workbook("E:\Employees.xlsx")

workbook.properties
<openpyxl.packaging.core.DocumentProperties object>
Parameters:
creator='openpyxl', title=None, description=None, subject=None, identifier=None, language=None, created=datetime.datetime(2015, 6, 5, 18, 17, 20), modified=datetime.datetime(2018, 10, 29, 10, 6, 33), lastModifiedBy='Mihai Catalin Teodosiu', category=None, contentStatus=None, version=None, revision=None, keywords=None, lastPrinted=None
workbook.sheetnames
['EmployeeData', 'Salaries', 'Skills']
workbook.active
<Worksheet "EmployeeData">
sheet = workbook['EmployeeData']

workbook.create_sheet('TestSheet')
<Worksheet "TestSheet">

workbook.save("E:\Employees.xlsx")
sheet = workbook['TestSheet']
workbook.remove(sheet)
# del workbook ['TestSheet']
workbook.save("E:\Employees.xlsx")

sheet = workbook['EmployeeData']

sheet.title
'EmployeeData'

dir(sheet)
['BREAK_COLUMN', 'BREAK_NONE', 'BREAK_ROW', 'HeaderFooter', 'ORIENTATION_LANDSCAPE', 'ORIENTATION_PORTRAIT', 'PAPERSIZE_A3', 'PAPERSIZE_A4', 'PAPERSIZE_A4_SMALL', 'PAPERSIZE_A5', 'PAPERSIZE_EXECUTIVE', 'PAPERSIZE_LEDGER', 'PAPERSIZE_LEGAL', 'PAPERSIZE_LETTER', 'PAPERSIZE_LETTER_SMALL', 'PAPERSIZE_STATEMENT', 'PAPERSIZE_TABLOID', 'SHEETSTATE_HIDDEN', 'SHEETSTATE_VERYHIDDEN', 'SHEETSTATE_VISIBLE', '_WorkbookChild__title', '__class__', '__delattr__', '__delitem__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__getitem__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__setitem__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__', '_add_cell', '_add_column', '_add_row', '_cells', '_cells_by_col', '_cells_by_row', '_charts', '_clean_merge_range', '_comments', '_current_row', '_default_title', '_drawing', '_get_cell', '_hyperlinks', '_id', '_images', '_invalid_row', '_move_cell', '_move_cells', '_parent', '_path', '_pivots', '_print_area', '_print_cols', '_print_rows', '_rel_type', '_rels', '_setup', '_tables', 'active_cell', 'add_chart', 'add_data_validation', 'add_image', 'add_pivot', 'add_table', 'append', 'auto_filter', 'calculate_dimension', 'cell', 'col_breaks', 'column_dimensions', 'columns', 'conditional_formatting', 'data_validations', 'delete_cols', 'delete_rows', 'dimensions', 'encoding', 'evenFooter', 'evenHeader', 'firstFooter', 'firstHeader', 'formula_attributes', 'freeze_panes', 'insert_cols', 'insert_rows', 'iter_cols', 'iter_rows', 'legacy_drawing', 'max_column', 'max_row', 'merge_cells', 'merged_cell_ranges', 'merged_cells', 'mime_type', 'min_column', 'min_row', 'move_range', 'oddFooter', 'oddHeader', 'orientation', 'page_breaks', 'page_margins', 'page_setup', 'paper_size', 'parent', 'path', 'print_area', 'print_options', 'print_title_cols', 'print_title_rows', 'print_titles', 'protection', 'row_breaks', 'row_dimensions', 'rows', 'scenarios', 'selected_cell', 'set_printer_settings', 'sheet_format', 'sheet_properties', 'sheet_state', 'sheet_view', 'show_gridlines', 'show_summary_below', 'show_summary_right', 'tables', 'title', 'unmerge_cells', 'values', 'views']






sheet.active_cell
'A1'

sheet.dimension
Traceback (most recent call last):
  File "<pyshell#29>", line 1, in <module>
    sheet.dimension
AttributeError: 'Worksheet' object has no attribute 'dimension'. Did you mean: 'dimensions'?
sheet.dimensions
'A1:F11'
sheet.sheet_format
<openpyxl.worksheet.dimensions.SheetFormatProperties object>
Parameters:
baseColWidth=8, defaultColWidth=None, defaultRowHeight=15.0, customHeight=None, zeroHeight=None, thickTop=None, thickBottom=None, outlineLevelRow=None, outlineLevelCol=0

sheet.sheet_properties
<openpyxl.worksheet.properties.WorksheetProperties object>
Parameters:
codeName=None, enableFormatConditionsCalculation=None, filterMode=None, published=None, syncHorizontal=None, syncRef=None, syncVertical=None, transitionEvaluation=None, transitionEntry=None, tabColor=None, outlinePr=<openpyxl.worksheet.properties.Outline object>
Parameters:
applyStyles=None, summaryBelow=True, summaryRight=True, showOutlineSymbols=None, pageSetUpPr=<openpyxl.worksheet.properties.PageSetupProperties object>
Parameters:
autoPageBreaks=None, fitToPage=None
sheet.sheet_state
'visible'
sheet.sheet_view
<openpyxl.worksheet.views.SheetView object>
Parameters:
windowProtection=None, showFormulas=None, showGridLines=None, showRowColHeaders=None, showZeros=None, rightToLeft=None, tabSelected=None, showRuler=None, showOutlineSymbols=None, defaultGridColor=None, showWhiteSpace=None, view=None, topLeftCell=None, colorId=None, zoomScale=250, zoomScaleNormal=250, zoomScaleSheetLayoutView=None, zoomScalePageLayoutView=None, zoomToFit=None, workbookViewId=0, pane=None, selection=[<openpyxl.worksheet.views.Selection object>
Parameters:
pane=None, activeCell='A1', activeCellId=None, sqref='A1']


sheet.max_row
11
sheet.min_row
1
sheet.max_coloum
Traceback (most recent call last):
  File "<pyshell#40>", line 1, in <module>
    sheet.max_coloum
AttributeError: 'Worksheet' object has no attribute 'max_coloum'. Did you mean: 'max_column'?
sheet.max_column
6

for i in sheet.values:
    print(i)

    
('ID', 'FirstName', 'LastName', 'Department', 'Phone', 'Address')
(1, 'Luke', 'Phillip', 'Sales', 1234567890, '1st Address, Miami')
(2, 'Jack', 'Darren', 'IT', 1234567891, '2nd Address, Miami')
(3, 'Ken', 'Wood', 'IT', 1234567892, '3rd Address, Miami')
(4, 'John', 'Wilson', 'Marketing', 1234567893, '4th Address, Miami')
(5, 'Emily', 'Larson', 'Marketing', 1234567894, '5th Address, Miami')
(6, 'Anna', 'Sullivan', 'Sales', 1234567895, '6th Address, Miami')
(7, 'Richard', 'Smith', 'Logistics', 1234567896, '7th Address, Miami')
(8, 'Ronnie', 'Moore', 'Sales', 1234567897, '8th Address, Miami')
(9, 'Test', 'Drake', 'IT', 1234567898, '9th Address, Miami')
(10, 'Wayne', 'Barker', 'Logistics', 1234567899, '10th Address, Miami')








































sheet = workbook ['EmployeeData']

sheet['B7'].value
'Anna'
sheet.cell(row = 6, column = 2).value
'Emily'

cell = sheet['B9']

cell.row
9

cel.column
Traceback (most recent call last):
  File "<pyshell#75>", line 1, in <module>
    cel.column
NameError: name 'cel' is not defined. Did you mean: 'cell'?
cell.column
2
cell.coordinate
'B9'
cell.data_type
's'
cell.encoding
'utf-8'
cell = sheet['B2']
cell.value = 'David'
workbook.save("E:\Employees.xlsx")
